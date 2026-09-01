from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def generate_xlsx(execution: dict, query: str, filename="Business_Report.xlsx"):
    """
    Generate an Excel workbook summarizing a completed workflow
    execution: a Summary sheet, a Step Log sheet, and a Final Report
    sheet with the response text.
    """

    workbook = Workbook()

    # ------------------------------------------------------------
    # SUMMARY SHEET
    # ------------------------------------------------------------
    summary = workbook.active
    summary.title = "Summary"

    summary["A1"] = "Enterprise Workflow Execution Report"
    summary["A1"].font = Font(bold=True, size=14)

    metrics = execution.get("metrics", {})
    rows = [
        ("Business Query", query),
        ("Status", execution.get("status", "")),
        ("Message", execution.get("message", "")),
        ("Run ID", execution.get("run_id", "")),
        ("Total Steps", metrics.get("total_steps", 0)),
        ("Completed Steps", metrics.get("completed_steps", 0)),
        ("Failed Steps", metrics.get("failed_steps", 0)),
        ("Success Rate (%)", metrics.get("success_rate", 0)),
        ("Total Duration (s)", metrics.get("total_duration_seconds", 0)),
    ]

    for i, (label, value) in enumerate(rows, start=3):
        summary[f"A{i}"] = label
        summary[f"A{i}"].font = Font(bold=True)
        summary[f"B{i}"] = str(value)

    for column, width in (("A", 24), ("B", 80)):
        summary.column_dimensions[column].width = width

    # ------------------------------------------------------------
    # STEP LOG SHEET
    # ------------------------------------------------------------
    step_sheet = workbook.create_sheet("Step Log")
    headers = ["Step", "Type", "Status", "Duration (s)", "Error"]

    for col_idx, header in enumerate(headers, start=1):
        cell = step_sheet.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    for row_idx, log in enumerate(metrics.get("step_logs", []), start=2):
        step_sheet.cell(row=row_idx, column=1, value=log.get("step", ""))
        step_sheet.cell(row=row_idx, column=2, value=log.get("type", ""))
        step_sheet.cell(row=row_idx, column=3, value=log.get("status", ""))
        step_sheet.cell(row=row_idx, column=4, value=log.get("duration_seconds", 0))
        step_sheet.cell(row=row_idx, column=5, value=log.get("error") or "")

    for col_idx, width in enumerate([20, 14, 12, 14, 60], start=1):
        step_sheet.column_dimensions[get_column_letter(col_idx)].width = width

    # ------------------------------------------------------------
    # FINAL REPORT SHEET
    # ------------------------------------------------------------
    report_sheet = workbook.create_sheet("Final Report")
    report_sheet["A1"] = "Final Business Response"
    report_sheet["A1"].font = Font(bold=True, size=12)

    final_response = execution.get("results", {}).get("response", "")

    row_idx = 3
    for line in str(final_response).split("\n"):
        cell = report_sheet.cell(row=row_idx, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True)
        row_idx += 1

    report_sheet.column_dimensions["A"].width = 100

    workbook.save(filename)

    return filename
